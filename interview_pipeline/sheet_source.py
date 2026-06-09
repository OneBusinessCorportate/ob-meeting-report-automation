"""Read the «Обучающий центр ОВ» candidate/interview table.

The source is a Google Sheet whose «Бух» tab lists accountant candidates. The
interview-transcript link lives in the «Первичн. собес. (ссылка на транскриб)»
column, but the sheet is hand-maintained and messy:
  * the link column is sometimes empty or the link sits in a neighbouring cell;
  * names embed phone numbers; contacts/emails are scattered;
  * the real transcript links we have are **Google Docs**, not Timeless.

So parsing is defensive: columns are matched by header *substring*, and any
http link to a known transcript host (docs.google.com / drive.google.com /
timeless) found anywhere in the row is treated as the transcript link.

Sources, in priority order (auto-detected unless INTERVIEW_SHEET_SOURCE is set):
  1. google_api  — Google Sheets API via a service account (most automatic);
  2. csv_url     — a published-to-web CSV export URL;
  3. local_xlsx  — a downloaded .xlsx file (openpyxl);
  4. local_csv   — a .csv file.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from meeting_pipeline.config import Config
from meeting_pipeline.utils import get_logger

log = get_logger("interview_pipeline.sheet")

# Hosts whose links we treat as a usable interview transcript / recording.
_TRANSCRIPT_HOSTS = ("docs.google.com", "drive.google.com", "timeless.day", "timeless")
_URL_RE = re.compile(r"https?://[^\s,;]+", re.IGNORECASE)
_EMAIL_RE = re.compile(r"[\w.\-+]+@[\w.\-]+\.\w+")
# Armenian / general phone-ish runs of digits, spaces, +, -, ().
_PHONE_RE = re.compile(r"(?:\+?\d[\d\-\s()]{6,}\d)")

# Map a sheet/tab name to a candidate "track".
_TRACK_BY_TAB = {
    "бух": "buh",
    "консультант бух": "consultant_buh",
    "юрист": "jurist",
}


@dataclass
class SheetCandidate:
    """One parsed candidate row (+ optional interview link)."""

    full_name: str
    track: str = "buh"
    role: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    contact_raw: Optional[str] = None
    resume_comment: Optional[str] = None
    sheet_status: Optional[str] = None
    grade_start: Optional[str] = None
    test_score: Optional[float] = None
    test_sent_at: Optional[str] = None       # ISO date
    probation_start: Optional[str] = None
    terminated_at: Optional[str] = None
    termination_reason: Optional[str] = None
    # Interview link (may be a Google Doc, a Timeless link, or empty).
    call_url: Optional[str] = None
    transcript_file: Optional[str] = None    # local MVP fallback (from a CSV column)
    source_sheet: Optional[str] = None
    source_row: Optional[int] = None
    source_column: Optional[str] = None      # which column the link came from
    metadata: Dict[str, Any] = field(default_factory=dict)


# --- low-level helpers --------------------------------------------------------

def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _to_iso_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # unparseable -> ignore the date rather than guess


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", ".").strip().rstrip("%")
    try:
        return float(text)
    except ValueError:
        return None


def _find_url(*cells: Any) -> Optional[str]:
    """Return the first transcript-host URL found in the given cells."""
    for cell in cells:
        if cell is None:
            continue
        for match in _URL_RE.findall(str(cell)):
            low = match.lower()
            if any(host in low for host in _TRANSCRIPT_HOSTS):
                return match.rstrip(".,);")
    # No known-host link: accept any URL as a last resort.
    for cell in cells:
        if cell is None:
            continue
        m = _URL_RE.search(str(cell))
        if m:
            return m.group(0).rstrip(".,);")
    return None


def _extract_contact(*cells: Any):
    blob = " ".join(str(c) for c in cells if c)
    email = _EMAIL_RE.search(blob)
    phone = _PHONE_RE.search(blob)
    return (
        email.group(0) if email else None,
        re.sub(r"\s+", " ", phone.group(0)).strip() if phone else None,
        blob.strip() or None,
    )


def _track_for(tab: str) -> str:
    return _TRACK_BY_TAB.get(_norm(tab), "other")


def _build_header_map(header: List[Any]) -> Dict[str, int]:
    """Map normalized header substrings to their column index (first wins)."""
    index: Dict[str, int] = {}
    for i, cell in enumerate(header):
        key = _norm(cell)
        if key and key not in index:
            index[key] = i
    return index


def _col(header_map: Dict[str, int], *needles: str) -> Optional[int]:
    for needle in needles:
        for key, idx in header_map.items():
            if needle in key:
                return idx
    return None


def _looks_like_header(row: List[Any]) -> bool:
    joined = _norm(" ".join(str(c) for c in row if c))
    return "претендент" in joined or "статус" in joined


def _row_to_candidate(
    row: List[Any], header_map: Dict[str, int], tab: str, row_number: int
) -> Optional[SheetCandidate]:
    def cell(idx: Optional[int]) -> Any:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    name_idx = _col(header_map, "претендент", "сотрудник", "кандидат") or 0
    name = str(cell(name_idx) or "").strip()
    if not name:
        return None

    resume_idx = _col(header_map, "резюме", "коммент", "контакт")
    status_idx = _col(header_map, "статус")
    test_date_idx = _col(header_map, "дата отправки теста", "отправки теста")
    test_res_idx = _col(header_map, "результаты теста", "результат теста")
    link_idx = _col(header_map, "ссылка на транскриб", "транскриб", "собес")
    grade_idx = _col(header_map, "грейд")
    probation_idx = _col(header_map, "испытательный")
    terminated_idx = _col(header_map, "прекращения сотрудничества", "прекращения")
    reason_idx = _col(header_map, "причина")
    role_idx = _col(header_map, "должность", "позиция")
    file_idx = _col(header_map, "transcript_file", "файл транскрипт", "файл")

    resume = cell(resume_idx)
    email, phone, contact_raw = _extract_contact(name, resume)
    # The name cell often carries a trailing phone — strip it for a clean name.
    clean_name = _PHONE_RE.sub("", name).replace("\n", " ").strip(" ,\t")
    clean_name = re.sub(r"\s+", " ", clean_name) or name.strip()

    # Link: prefer the designated column, else scan the whole row.
    link = _find_url(cell(link_idx))
    if not link:
        link = _find_url(*row)
    source_column = None
    if link and link_idx is not None and _find_url(cell(link_idx)):
        source_column = "Первичн. собес. (ссылка на транскриб)"
    elif link:
        source_column = "scanned_row"

    return SheetCandidate(
        full_name=clean_name,
        track=_track_for(tab),
        role=str(cell(role_idx)).strip() if cell(role_idx) else None,
        email=email,
        phone=phone,
        contact_raw=contact_raw if (email or phone) else (str(resume).strip() if resume else None),
        resume_comment=str(resume).strip() if resume else None,
        sheet_status=str(cell(status_idx)).strip() if cell(status_idx) else None,
        grade_start=str(cell(grade_idx)).strip() if cell(grade_idx) else None,
        test_score=_to_float(cell(test_res_idx)),
        test_sent_at=_to_iso_date(cell(test_date_idx)),
        probation_start=_to_iso_date(cell(probation_idx)),
        terminated_at=_to_iso_date(cell(terminated_idx)),
        termination_reason=str(cell(reason_idx)).strip() if cell(reason_idx) else None,
        call_url=link,
        transcript_file=str(cell(file_idx)).strip() if cell(file_idx) else None,
        source_sheet=tab,
        source_row=row_number,
        source_column=source_column,
    )


def _parse_rows(rows: List[List[Any]], tab: str) -> List[SheetCandidate]:
    """Turn a matrix of cells (incl. header) into candidates."""
    if not rows:
        return []
    # Find the header row (first row that looks like headers, else row 0).
    header_row = 0
    for i, row in enumerate(rows[:5]):
        if _looks_like_header(row):
            header_row = i
            break
    header_map = _build_header_map(rows[header_row])
    out: List[SheetCandidate] = []
    for i, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        if not any(str(c).strip() for c in row if c is not None):
            continue  # blank row
        cand = _row_to_candidate(list(row), header_map, tab, i)
        if cand:
            out.append(cand)
    log.info("Parsed %d candidate row(s) from tab '%s'", len(out), tab)
    return out


# --- source loaders -----------------------------------------------------------

def from_xlsx(path: str, tabs: List[str]) -> List[SheetCandidate]:
    import openpyxl  # lazy; only needed for local xlsx

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: List[SheetCandidate] = []
    for tab in tabs:
        if tab not in wb.sheetnames:
            log.warning("Tab '%s' not found in %s (have: %s)", tab, path, wb.sheetnames)
            continue
        ws = wb[tab]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        out.extend(_parse_rows(rows, tab))
    return out


def from_csv_text(text: str, tab: str) -> List[SheetCandidate]:
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    return _parse_rows(rows, tab)


def from_csv_file(path: str, tab: str = "Бух") -> List[SheetCandidate]:
    return from_csv_text(Path(path).read_text(encoding="utf-8"), tab)


def from_csv_url(url: str, tab: str = "Бух", session: Any = None) -> List[SheetCandidate]:
    if session is None:
        import requests  # lazy

        session = requests.Session()
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    return from_csv_text(resp.text, tab)


def from_google_api(
    spreadsheet_id: str, tabs: List[str], config: Config
) -> List[SheetCandidate]:
    """Read tabs via the Google Sheets API using a service account."""
    from .google_creds import build_gspread_client  # lazy

    gc = build_gspread_client(config)
    sh = gc.open_by_key(spreadsheet_id)
    out: List[SheetCandidate] = []
    for tab in tabs:
        try:
            ws = sh.worksheet(tab)
        except Exception as exc:  # worksheet missing / no access
            log.warning("Cannot open tab '%s': %s", tab, exc)
            continue
        rows = ws.get_all_values()
        out.extend(_parse_rows(rows, tab))
    return out


def load_candidates(
    config: Config,
    *,
    xlsx_path: Optional[str] = None,
    csv_path: Optional[str] = None,
    tabs: Optional[List[str]] = None,
    session: Any = None,
) -> List[SheetCandidate]:
    """Resolve the candidate table from the best available source.

    Explicit CLI args (xlsx_path / csv_path) win. Otherwise auto-detect from
    config: Google API → published CSV URL → local xlsx.
    """
    tab_list = tabs or [t.strip() for t in (config.interview_sheet_tabs or "Бух").split(",") if t.strip()]

    if xlsx_path:
        return from_xlsx(xlsx_path, tab_list)
    if csv_path:
        return from_csv_file(csv_path, tab_list[0] if tab_list else "Бух")

    mode = (config.interview_sheet_source or "").strip().lower()

    if (mode in ("", "google_api")) and config.interview_spreadsheet_id and (
        config.google_service_account_json or config.google_service_account_file
    ):
        log.info("Sheet source: Google Sheets API (spreadsheet %s)", config.interview_spreadsheet_id)
        return from_google_api(config.interview_spreadsheet_id, tab_list, config)

    if (mode in ("", "csv_url")) and config.interview_sheet_csv_url:
        log.info("Sheet source: published CSV URL")
        return from_csv_url(config.interview_sheet_csv_url, tab_list[0] if tab_list else "Бух", session)

    if (mode in ("", "local_xlsx")) and config.interview_local_xlsx:
        log.info("Sheet source: local xlsx %s", config.interview_local_xlsx)
        return from_xlsx(config.interview_local_xlsx, tab_list)

    log.warning(
        "No sheet source configured. Pass --xlsx/--csv, or set INTERVIEW_SPREADSHEET_ID "
        "+ Google creds, INTERVIEW_SHEET_CSV_URL, or INTERVIEW_LOCAL_XLSX."
    )
    return []
